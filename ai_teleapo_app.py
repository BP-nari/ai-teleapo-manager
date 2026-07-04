import streamlit as st
import pandas as pd
import re
import csv
import datetime
import openpyxl
import hashlib
import json
import time
import pickle
from pathlib import Path
from io import BytesIO, StringIO

st.set_page_config(
    page_title="AI管理システム",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: 'Helvetica Neue', Arial, 'Hiragino Kaku Gothic ProN', Meiryo, sans-serif;
}
.stApp { background-color: #f5f6fa; }
.main .block-container { padding: 2rem 2.5rem; max-width: 1200px; }

/* サイドバー */
section[data-testid="stSidebar"] { background: #0f172a !important; }
section[data-testid="stSidebar"] * { color: #cbd5e1 !important; }
section[data-testid="stSidebar"] .stButton > button {
    background: transparent !important; border: none !important;
    color: #94a3b8 !important; text-align: left !important;
    padding: 0.5rem 1rem !important; border-radius: 6px !important;
    font-size: 0.875rem !important; font-weight: 500 !important;
    width: 100% !important;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    background: #1e293b !important; color: #f1f5f9 !important;
}
.sb-logo { padding: 1.5rem 1rem 1rem 1rem; border-bottom: 1px solid #1e293b; margin-bottom: 0.5rem; }
.sb-logo h2 { font-size: 0.95rem; font-weight: 800; color: #f1f5f9 !important; margin: 0; }
.sb-logo p  { font-size: 0.72rem; color: #64748b !important; margin: 0.2rem 0 0 0; }
.sb-section {
    font-size: 0.65rem; font-weight: 700; letter-spacing: 0.1em;
    text-transform: uppercase; color: #475569 !important;
    padding: 0.9rem 1rem 0.3rem 1rem; margin: 0;
}

/* ページヘッダー */
.topbar-ds {
    background: #1e3a8a; border-radius: 10px;
    padding: 1.25rem 1.75rem; margin-bottom: 1.5rem; color: white;
}
.topbar-ai {
    background: #0f766e; border-radius: 10px;
    padding: 1.25rem 1.75rem; margin-bottom: 1.5rem; color: white;
}
.topbar-hist {
    background: #1e293b; border-radius: 10px;
    padding: 1.25rem 1.75rem; margin-bottom: 1.5rem; color: white;
}
.topbar-ds h2, .topbar-ai h2, .topbar-hist h2 {
    font-size: 1.35rem; font-weight: 700; margin: 0 0 0.2rem 0;
}
.topbar-ds p, .topbar-ai p, .topbar-hist p {
    font-size: 0.82rem; opacity: 0.8; margin: 0;
}

/* カード */
.card {
    background: white; border-radius: 10px; padding: 1.5rem;
    margin-bottom: 1rem; border: 1px solid #e2e8f0;
    box-shadow: 0 1px 4px rgba(0,0,0,0.04);
}
.card-label {
    font-size: 0.75rem; font-weight: 700; letter-spacing: 0.06em;
    text-transform: uppercase; color: #64748b; margin-bottom: 0.8rem;
}

/* ステップ番号 */
.sn-ds {
    display: inline-flex; align-items: center; justify-content: center;
    width: 22px; height: 22px; border-radius: 50%;
    background: #1e3a8a; color: white;
    font-size: 0.72rem; font-weight: 700;
    margin-right: 0.5rem; vertical-align: middle;
}
.sn-ai {
    display: inline-flex; align-items: center; justify-content: center;
    width: 22px; height: 22px; border-radius: 50%;
    background: #0f766e; color: white;
    font-size: 0.72rem; font-weight: 700;
    margin-right: 0.5rem; vertical-align: middle;
}

/* アラート */
.alert-ok   { background:#f0fdf4; border-left:4px solid #22c55e; border-radius:6px; padding:0.9rem 1.1rem; margin:0.8rem 0; font-size:0.875rem; color:#166534; }
.alert-info { background:#eff6ff; border-left:4px solid #3b82f6; border-radius:6px; padding:0.9rem 1.1rem; margin:0.8rem 0; font-size:0.875rem; color:#1e40af; }
.alert-warn { background:#fffbeb; border-left:4px solid #f59e0b; border-radius:6px; padding:0.9rem 1.1rem; margin:0.8rem 0; font-size:0.875rem; color:#92400e; }

/* ホームカード */
.home-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; margin-top: 1.5rem; }
.home-card { border-radius: 10px; padding: 1.5rem; color: white; position: relative; }
.home-card-ds  { background: #1e3a8a; }
.home-card-ai  { background: #0f766e; }
.home-card-wip { background: #475569; }
.home-card .tag {
    position: absolute; top: 1rem; right: 1rem;
    background: rgba(255,255,255,0.2); border-radius: 4px;
    padding: 0.15rem 0.55rem; font-size: 0.65rem; font-weight: 700; letter-spacing: 0.08em;
}
.home-card h3 { font-size: 1rem; font-weight: 700; margin: 0.5rem 0 0.4rem 0; }
.home-card p  { font-size: 0.8rem; opacity: 0.85; margin: 0; line-height: 1.6; }

/* 工事中 */
.wip-block {
    background: #1e293b; border-radius: 10px; padding: 5rem 2rem;
    text-align: center; color: #94a3b8; margin: 2rem 0;
}
.wip-block h2 { font-size: 1.5rem; font-weight: 700; color: #e2e8f0; margin: 0 0 0.6rem 0; }
.wip-block p  { font-size: 0.9rem; line-height: 1.8; margin: 0; }

/* ジョブカード */
.job-card {
    background: white; border-radius: 8px; padding: 1.2rem 1.4rem;
    margin-bottom: 0.75rem; border: 1px solid #e2e8f0;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04);
}
.job-card-header {
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 0.75rem; padding-bottom: 0.6rem; border-bottom: 1px solid #f1f5f9;
}
.job-id { font-family: monospace; font-size: 0.85rem; font-weight: 700; color: #1e293b; }
.job-date { font-size: 0.78rem; color: #94a3b8; }
.job-meta { display: flex; gap: 1.5rem; flex-wrap: wrap; }
.job-meta-item { font-size: 0.82rem; color: #475569; }
.job-meta-item strong { color: #0f172a; }
.job-fname {
    font-size: 0.82rem; color: #334155; margin-top: 0.5rem;
    background: #f8fafc; border-radius: 4px; padding: 0.35rem 0.7rem;
    font-family: monospace; cursor: text; border: 1px solid #e2e8f0;
    word-break: break-all;
}
.badge-ds { background: #dbeafe; color: #1e40af; border-radius: 4px; padding: 0.15rem 0.5rem; font-size: 0.68rem; font-weight: 700; letter-spacing: 0.06em; }
.badge-ai { background: #ccfbf1; color: #0f766e; border-radius: 4px; padding: 0.15rem 0.5rem; font-size: 0.68rem; font-weight: 700; letter-spacing: 0.06em; }

/* ダウンロードボタン */
.stDownloadButton > button {
    background: #16a34a !important; color: white !important;
    border: none !important; border-radius: 7px !important;
    font-weight: 700 !important; font-size: 0.9rem !important;
    padding: 0.65rem 1.8rem !important; width: 100% !important;
    margin-top: 0.5rem !important;
}
.stDownloadButton > button:hover { background: #15803d !important; }

/* テーブル */
.stat-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
.stat-table th {
    background: #f1f5f9; color: #475569; padding: 0.5rem 0.75rem;
    text-align: left; font-weight: 600; font-size: 0.78rem; border-bottom: 1px solid #e2e8f0;
}
.stat-table td { padding: 0.45rem 0.75rem; border-bottom: 1px solid #f1f5f9; color: #1e293b; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# ジョブ履歴管理
# ============================================================
HISTORY_FILE = Path("job_history.json")

def load_jobs():
    try:
        if HISTORY_FILE.exists():
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return []

def save_jobs(jobs):
    try:
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(jobs, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def add_job(job_type, filename, rows_in, rows_out):
    jobs = load_jobs()
    job_id = datetime.datetime.now().strftime("%Y%m%d_%H%M") + "_" + hashlib.md5(str(time.time()).encode()).hexdigest()[:5].upper()
    jobs.insert(0, {
        "job_id":   job_id,
        "type":     job_type,
        "filename": filename,
        "rows_in":  rows_in,
        "rows_out": rows_out,
        "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })
    save_jobs(jobs[:200])  # 最大200件保持
    return job_id


# ============================================================
# ロジック: DS リスト成形（Excel入力）
# ============================================================
COL_ID       = 3
COL_CUSTOMER = 7
COL_ADDRESS  = 8
COL_PHONE    = 9
MAX_KEEP_BEFORE_O = 14

def normalize_phone_ds(value):
    if value is None:
        return ''
    s = str(value).strip()
    if re.fullmatch(r'\d+\.0', s):
        s = s[:-2]
    digits = re.sub(r'[^0-9]', '', s)
    if len(digits) == 10 and digits[0] in ('7', '8', '9'):
        digits = '0' + digits
    return digits

def is_valid_phone_ds(value):
    digits = normalize_phone_ds(value)
    if not digits.isdigit():
        return False
    if len(digits) not in (10, 11):
        return False
    if not digits.startswith('0'):
        return False
    return True

def safe_str_ds(value):
    if value is None:
        return ''
    if isinstance(value, datetime.datetime):
        return value.strftime('%Y/%m/%d')
    if isinstance(value, datetime.date):
        return value.strftime('%Y/%m/%d')
    if isinstance(value, datetime.timedelta):
        total = int(value.total_seconds())
        h, rem = divmod(total, 3600)
        m, s = divmod(rem, 60)
        return f'{h}:{m:02d}:{s:02d}'
    return str(value).strip()

def process_ds_list(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    priority_indices = [COL_CUSTOMER, COL_PHONE, COL_ID, COL_ADDRESS]
    priority_headers = ['顧客名', '電話番号', 'IDの頭にID', '住所統合']
    remaining_indices = [
        i for i in range(min(MAX_KEEP_BEFORE_O, len(headers)))
        if i not in set(priority_indices)
    ]
    out_indices = priority_indices + remaining_indices
    out_headers = priority_headers + [safe_str_ds(headers[i]) for i in remaining_indices]

    removed_rows = []
    kept_rows = [out_headers]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None or str(cell).strip() == '' for cell in row):
            removed_rows.append((row_idx, '空白行'))
            continue
        customer_name = row[COL_CUSTOMER] if len(row) > COL_CUSTOMER else None
        if customer_name is None or str(customer_name).strip() == '':
            removed_rows.append((row_idx, '顧客名が未入力'))
            continue
        phone_raw = row[COL_PHONE] if len(row) > COL_PHONE else None
        if not is_valid_phone_ds(phone_raw):
            removed_rows.append((row_idx, f'電話番号「{phone_raw}」が不正形式'))
            continue
        out_row = []
        for idx in out_indices:
            value = row[idx] if len(row) > idx else None
            if idx == COL_PHONE:
                out_row.append(normalize_phone_ds(value))
            else:
                out_row.append(safe_str_ds(value))
        kept_rows.append(out_row)

    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerows(kept_rows)
    csv_bytes = buf.getvalue().encode('utf-8-sig')
    return csv_bytes, kept_rows, removed_rows, ws.max_row


# ============================================================
# ロジック: DS FM返送（CSV入力）
# ============================================================
CONVERSION_MAP = {
    "AI終話": "NG", "応答なし": "留守", "AI対応不可": "留守電",
    "AI同時接続": "再コール・転送成功", "留守番電話": "留守電",
    "コール結果記録中": "再コール・転送成功", "手動受信_未対応": "受電あり",
    "トスアップ応答前終了": "再コール・転送成功", "コール結果未登録": "再コール・転送成功",
    "転送\u3000NG": "NG", "転送\u3000再コール": "再コール・転送成功",
    "転送\u3000アポ禁": "アポ禁", "転送\u3000留守電": "留守電",
    "受電\u3000NG": "NG", "受電\u3000再コール": "再コール",
    "受電\u3000アポ禁": "アポ禁", "受電\u3000電話APO": "受電電話APO",
    "受電\u3000留守電": "留守電", "AIコールNG": "NG", "AIホットリード": "再コール・転送成功",
}
DROP_FLAGS = {"転送　お題成立", "転送　紐づけ", "受電　お題成立", "受電　紐づけ"}

# FM返送の出力列定義（この順序・列名で出力する）
FM_OUTPUT_COLUMNS = [
    "会社名", "電話番号", "前回結果", "前回コール日", "コール時間", "履歴内容", "営業担当", "IDの頭にID",
]

def process_ds_fm(raw_bytes):
    df = None
    used_enc = None
    for enc in ["utf-8-sig", "utf-8", "cp932", "shift_jis"]:
        try:
            df = pd.read_csv(BytesIO(raw_bytes), encoding=enc)
            used_enc = enc
            break
        except Exception:
            continue
    if df is None:
        raise ValueError("CSVを読み込めませんでした。文字コードを確認してください。")

    rows_before = len(df)

    # コール日時 → 前回コール日（YYYY/M/D）+ コール時間（HH:MM:SS）に分割
    if "コール日時" in df.columns:
        call_datetime = pd.to_datetime(df["コール日時"], errors="coerce")
        df["前回コール日"] = call_datetime.apply(
            lambda dt: f"{dt.year}/{dt.month}/{dt.day}" if pd.notna(dt) else ""
        )
        df["コール時間"] = call_datetime.dt.strftime("%H:%M:%S")
        df = df.drop(columns=["コール日時"])

    # コール結果 → 前回結果 に変換・列名変更
    if "コール結果" in df.columns:
        original_result = df["コール結果"].astype("string").str.strip().str.replace(" ", "　", regex=False)
        df = df[~original_result.isin(DROP_FLAGS)].copy()
        original_result = df["コール結果"].astype("string").str.strip().str.replace(" ", "　", regex=False)
        df["前回結果"] = original_result.map(CONVERSION_MAP).fillna(original_result)
        df = df.drop(columns=["コール結果"])

    # 名前 / 顧客名 → 会社名 に列名変更
    if "名前" in df.columns and "会社名" not in df.columns:
        df = df.rename(columns={"名前": "会社名"})
    elif "顧客名" in df.columns and "会社名" not in df.columns:
        df = df.rename(columns={"顧客名": "会社名"})

    # コールメモ → 履歴内容 に列名変更
    if "コールメモ" in df.columns and "履歴内容" not in df.columns:
        df = df.rename(columns={"コールメモ": "履歴内容"})

    # 担当者 / コール担当者 → 営業担当 に列名変更
    if "担当者" in df.columns and "営業担当" not in df.columns:
        df = df.rename(columns={"担当者": "営業担当"})
    elif "コール担当者" in df.columns and "営業担当" not in df.columns:
        df = df.rename(columns={"コール担当者": "営業担当"})

    # 出力列のみを指定順で抽出（存在しない列は空列として補完）
    for col in FM_OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[FM_OUTPUT_COLUMNS]

    rows_after = len(df)
    csv_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    return csv_bytes, df, rows_before, rows_after, used_enc


# ============================================================
# ロジック: AI FM返送（CSV入力 → Excel出力 1シート）
# ============================================================
AI_FM_OUTPUT_COLUMNS = [
    "日時", "種別", "会社名", "担当者名", "電話番号",
    "ステータス", "最終ステータス", "ラベル", "通話時間",
    "架電結果", "要約", "詳細リンク",
    "再電話回数", "メールアドレス", "最終試行", "総試行回数",
    "ソース", "IDの頭にID",
]

def process_ai_fm(raw_bytes, source_filename=""):
    df = None
    used_enc = None
    for enc in ["utf-8-sig", "utf-8", "cp932", "shift_jis"]:
        try:
            df = pd.read_csv(BytesIO(raw_bytes), encoding=enc)
            used_enc = enc
            break
        except Exception:
            continue
    if df is None:
        raise ValueError("CSVを読み込めませんでした。文字コードを確認してください。")

    rows_before = len(df)

    # IDの頭にID で重複排除（最新コールのみ残す）
    if "IDの頭にID" in df.columns and "日時" in df.columns:
        df["_dt"] = pd.to_datetime(df["日時"], errors="coerce")
        df = df.sort_values("_dt", ascending=False)
        df = df.drop_duplicates(subset=["IDの頭にID"], keep="first")
        df = df.drop(columns=["_dt"])
        df = df.sort_values("日時", ascending=False).reset_index(drop=True)

    rows_after = len(df)

    # ソース列を付与
    df["ソース"] = source_filename

    # 出力列を固定（存在しない列は空列で補完）
    for col in AI_FM_OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[AI_FM_OUTPUT_COLUMNS]

    # Excelとして出力（1シート）
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="通話履歴")
    excel_bytes = buf.getvalue()

    return excel_bytes, df, rows_before, rows_after, used_enc


# ============================================================
# ロジック: AI リスト成形（Excel入力）
# ============================================================
def process_ai_list(file_bytes):
    df = pd.read_excel(BytesIO(file_bytes))
    upload_df = df.copy()

    if '顧客名' in upload_df.columns:
        upload_df = upload_df.rename(columns={'顧客名': '社名'})

    required_columns = ['社名', '電話番号', '住所統合']
    available_columns = [col for col in required_columns if col in upload_df.columns]
    if available_columns:
        upload_df = upload_df[available_columns].copy()

    if '社名' in upload_df.columns:
        upload_df['社名'] = upload_df['社名'].astype(str).str[:50]

    # 「IDの頭にID」→「ID」に変換して付与
    if 'IDの頭にID' in df.columns:
        upload_df['ID'] = df['IDの頭にID'].values

    try:
        csv_bytes = upload_df.to_csv(index=False, encoding='shift_jis').encode('shift_jis')
    except UnicodeEncodeError:
        csv_bytes = upload_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')

    return csv_bytes, upload_df, len(df)


# ============================================================
# セッション初期化
# ============================================================
def init_session():
    if 'page' not in st.session_state:
        st.session_state.page = 'home'


# ============================================================
# サイドバー
# ============================================================
def render_sidebar():
    with st.sidebar:
        st.markdown("""
        <div class="sb-logo">
            <h2>AI管理システム</h2>
            <p>架電リスト成形 / FM返送</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<p class="sb-section">Dial Shift</p>', unsafe_allow_html=True)
        if st.button("リスト成形", key="nav_ds_list", use_container_width=True):
            st.session_state.page = 'ds_list'; st.rerun()
        if st.button("ファイルメーカー返送", key="nav_ds_fm", use_container_width=True):
            st.session_state.page = 'ds_fm'; st.rerun()

        st.markdown('<p class="sb-section">AI Teleapo</p>', unsafe_allow_html=True)
        if st.button("リスト成形", key="nav_ai_list", use_container_width=True):
            st.session_state.page = 'ai_list'; st.rerun()
        if st.button("ファイルメーカー返送", key="nav_ai_fm", use_container_width=True):
            st.session_state.page = 'ai_fm'; st.rerun()

        st.markdown('<p class="sb-section">System</p>', unsafe_allow_html=True)
        if st.button("ジョブ履歴", key="nav_history", use_container_width=True):
            st.session_state.page = 'history'; st.rerun()
        if st.button("ホーム", key="nav_home", use_container_width=True):
            st.session_state.page = 'home'; st.rerun()


# ============================================================
# ページ: ホーム
# ============================================================
def page_home():
    st.markdown("""
    <div style="padding: 2rem 0 1rem 0;">
        <h1 style="font-size:1.8rem;font-weight:800;color:#0f172a;margin:0 0 0.3rem 0;letter-spacing:-0.02em;">AI管理システム</h1>
        <p style="color:#64748b;font-size:0.9rem;margin:0;">ダイヤルシフト・AIテレアポのリスト成形とファイルメーカー返送</p>
    </div>
    """, unsafe_allow_html=True)

    jobs = load_jobs()
    total = len(jobs)
    ds_count = sum(1 for j in jobs if j.get('type', '').startswith('DS'))
    ai_count = sum(1 for j in jobs if j.get('type', '').startswith('AI'))

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("総ジョブ数", f"{total:,}")
    with c2:
        st.metric("ダイヤルシフト", f"{ds_count:,}")
    with c3:
        st.metric("AIテレアポ", f"{ai_count:,}")

    # カードをボタンで実装（クリックでページ遷移）
    st.markdown("""
    <style>
    div[data-testid="column"] .stButton > button {
        width: 100% !important;
        height: 160px !important;
        border-radius: 10px !important;
        border: none !important;
        color: white !important;
        font-size: 1rem !important;
        font-weight: 700 !important;
        text-align: left !important;
        padding: 1.4rem 1.5rem !important;
        line-height: 1.6 !important;
        white-space: pre-wrap !important;
        cursor: pointer !important;
        transition: opacity 0.15s !important;
    }
    div[data-testid="column"] .stButton > button:hover {
        opacity: 0.88 !important;
        transform: translateY(-2px) !important;
    }
    button[kind="secondary"][data-testid="baseButton-secondary"]:nth-child(1) { background: #1e3a8a !important; }
    </style>
    """, unsafe_allow_html=True)

    # ホームカード用CSS（タイトルクリックで遷移するボタン風カード）
    st.markdown("""
    <style>
    .home-nav-btn > button {
        width: 100% !important; height: auto !important; min-height: 140px !important;
        border-radius: 10px !important; border: none !important;
        color: white !important; font-size: 0.85rem !important;
        text-align: left !important; padding: 1.4rem 1.5rem !important;
        line-height: 1.6 !important; white-space: pre-wrap !important;
        cursor: pointer !important; transition: opacity 0.15s, transform 0.15s !important;
    }
    .home-nav-btn > button:hover { opacity: 0.88 !important; transform: translateY(-2px) !important; }
    .home-nav-ds > button  { background: #1e3a8a !important; }
    .home-nav-ai > button  { background: #0f766e !important; }
    </style>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    col3, col4 = st.columns(2)

    with col1:
        st.markdown('<div class="home-nav-btn home-nav-ds">', unsafe_allow_html=True)
        if st.button(
            "DIAL SHIFT  ·  01\n\nリスト成形\n\nExcelをアップロードして顧客名・電話番号・住所を整形。電話番号バリデーション付き。",
            key="home_ds_list", use_container_width=True
        ):
            st.session_state.page = 'ds_list'; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        st.markdown('<div class="home-nav-btn home-nav-ds">', unsafe_allow_html=True)
        if st.button(
            "DIAL SHIFT  ·  02\n\nファイルメーカー返送\n\nダイヤルシフトの履歴CSVをアップロードしてコール結果変換・列名整形してCSV出力。",
            key="home_ds_fm", use_container_width=True
        ):
            st.session_state.page = 'ds_fm'; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    with col3:
        st.markdown('<div class="home-nav-btn home-nav-ai">', unsafe_allow_html=True)
        if st.button(
            "AI TELEAPO  ·  03\n\nリスト成形\n\nFileMakerデータをAIテレアポ投入用CSVに変換。社名・電話番号・住所統合・IDを整形。",
            key="home_ai_list", use_container_width=True
        ):
            st.session_state.page = 'ai_list'; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    with col4:
        st.markdown('<div class="home-nav-btn home-nav-ai">', unsafe_allow_html=True)
        if st.button(
            "AI TELEAPO  ·  04\n\nファイルメーカー返送\n\nAIテレアポの通話履歴CSVをアップロードして重複排除・列整形しExcel出力。",
            key="home_ai_fm", use_container_width=True
        ):
            st.session_state.page = 'ai_fm'; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)


# ============================================================
# ページ: DS リスト成形
# ============================================================
def page_ds_list():
    st.markdown("""
    <div class="topbar-ds">
        <h2>ダイヤルシフト — リスト成形</h2>
        <p>Excelをアップロードして顧客名・電話番号・ID・住所を先頭に並び替え、電話番号バリデーション後にCSV出力</p>
    </div>
    """, unsafe_allow_html=True)

    col_main, col_side = st.columns([3, 1])

    with col_main:
        st.markdown('<div class="card"><div class="card-label"><span class="sn-ds">1</span>Excelファイルをアップロード</div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("FileMakerからエクスポートしたExcel (.xlsx / .xls) を選択",
                                    type=['xlsx', 'xls'], key="ds_list_upload", label_visibility="collapsed")
        st.markdown('</div>', unsafe_allow_html=True)

        if uploaded:
            file_bytes = uploaded.read()
            st.markdown(f'<div class="alert-ok">読み込み完了 — {uploaded.name} ({len(file_bytes)//1024:,} KB)</div>', unsafe_allow_html=True)

            st.markdown('<div class="card"><div class="card-label"><span class="sn-ds">2</span>成形を実行</div>', unsafe_allow_html=True)
            if st.button("成形を実行", type="primary", key="ds_list_run"):
                with st.spinner("処理中..."):
                    try:
                        csv_bytes, kept_rows, removed_rows, total_rows = process_ds_list(file_bytes)
                        data_in  = total_rows - 1
                        data_out = len(kept_rows) - 1
                        deleted  = len(removed_rows)

                        st.markdown(f'<div class="alert-ok">成形完了 — 入力: {data_in:,} 件 / 出力: {data_out:,} 件 / 削除: {deleted:,} 件</div>', unsafe_allow_html=True)

                        job_id = add_job("DS リスト成形", uploaded.name, data_in, data_out)
                        st.markdown(f'<div class="alert-info">ジョブID: <code>{job_id}</code></div>', unsafe_allow_html=True)

                        if removed_rows:
                            with st.expander(f"削除された行の詳細 ({deleted} 件)"):
                                del_df = pd.DataFrame(removed_rows, columns=["行番号", "削除理由"])
                                st.dataframe(del_df, use_container_width=True, hide_index=True)

                        if len(kept_rows) > 1:
                            with st.expander("出力データプレビュー（先頭5件）"):
                                st.dataframe(pd.DataFrame(kept_rows[1:6], columns=kept_rows[0]),
                                             use_container_width=True, hide_index=True)

                        base = uploaded.name.rsplit('.', 1)[0]
                        date_str = datetime.datetime.now().strftime("%Y%m%d")
                        st.markdown("---")
                        st.download_button("CSVをダウンロード", data=csv_bytes,
                                           file_name=f"{base}_{date_str}_整形済.csv",
                                           mime="text/csv", key="ds_list_dl")
                    except Exception as e:
                        st.error(f"処理エラー: {e}")
            st.markdown('</div>', unsafe_allow_html=True)

    with col_side:
        st.markdown("""
        <div class="card">
            <div class="card-label">処理内容</div>
            <table class="stat-table">
                <tr><td>顧客名・電話番号・ID・住所を先頭に移動</td></tr>
                <tr><td>電話番号バリデーション（0始まり 10〜11桁）</td></tr>
                <tr><td>空白行・顧客名未入力行を削除</td></tr>
                <tr><td>A〜N列（14列）を保持</td></tr>
                <tr><td>UTF-8 BOM付きCSVで出力</td></tr>
            </table>
        </div>
        """, unsafe_allow_html=True)


# ============================================================
# ページ: DS FM返送
# ============================================================
def page_ds_fm():
    st.markdown("""
    <div class="topbar-ds">
        <h2>ダイヤルシフト — ファイルメーカー返送</h2>
        <p>ダイヤルシフトの履歴CSVをアップロードしてコール結果変換・有効無効判定を付与してCSV出力</p>
    </div>
    """, unsafe_allow_html=True)

    col_main, col_side = st.columns([3, 1])

    with col_main:
        st.markdown('<div class="card"><div class="card-label"><span class="sn-ds">1</span>ダイヤルシフト履歴CSVをアップロード</div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("ダイヤルシフトからエクスポートしたCSVを選択",
                                    type=['csv'], key="ds_fm_upload", label_visibility="collapsed")
        st.markdown('</div>', unsafe_allow_html=True)

        if uploaded:
            raw = uploaded.read()
            st.markdown(f'<div class="alert-ok">読み込み完了 — {uploaded.name} ({len(raw)//1024:,} KB)</div>', unsafe_allow_html=True)

            st.markdown('<div class="card"><div class="card-label"><span class="sn-ds">2</span>変換を実行</div>', unsafe_allow_html=True)
            if st.button("変換を実行", type="primary", key="ds_fm_run"):
                with st.spinner("処理中..."):
                    try:
                        csv_bytes, result_df, rows_before, rows_after, used_enc = process_ds_fm(raw)

                        st.markdown(f'<div class="alert-ok">変換完了 — 入力: {rows_before:,} 件 / 出力: {rows_after:,} 件 / 削除: {rows_before - rows_after:,} 件 &nbsp;|&nbsp; 文字コード: {used_enc}</div>', unsafe_allow_html=True)

                        job_id = add_job("DS FM返送", uploaded.name, rows_before, rows_after)
                        st.markdown(f'<div class="alert-info">ジョブID: <code>{job_id}</code></div>', unsafe_allow_html=True)

                        if "前回結果" in result_df.columns:
                            with st.expander("前回結果の分布"):
                                dist = result_df["前回結果"].value_counts().reset_index()
                                dist.columns = ["前回結果", "件数"]
                                st.dataframe(dist, use_container_width=True, hide_index=True)

                        with st.expander("出力データプレビュー（先頭5件）"):
                            st.dataframe(result_df.head(5), use_container_width=True)

                        base = uploaded.name.rsplit('.', 1)[0]
                        date_str = datetime.datetime.now().strftime("%Y%m%d")
                        st.markdown("---")
                        st.download_button("CSVをダウンロード", data=csv_bytes,
                                           file_name=f"{base}_{date_str}_formatted.csv",
                                           mime="text/csv", key="ds_fm_dl")
                    except Exception as e:
                        st.error(f"処理エラー: {e}")
            st.markdown('</div>', unsafe_allow_html=True)

    with col_side:
        st.markdown("""
        <div class="card">
            <div class="card-label">処理内容</div>
            <table class="stat-table">
                <tr><td>顧客名 → 会社名 に列名変更</td></tr>
                <tr><td>コール日時 → 前回コール日（YYYY/M/D）/ コール時間 に分割</td></tr>
                <tr><td>お題成立・紐づけ行を削除</td></tr>
                <tr><td>コール結果 → 前回結果 に変換・列名変更</td></tr>
                <tr><td>コール担当者 → 営業担当 に列名変更</td></tr>
                <tr><td>出力列を固定（8列）しUTF-8 BOM付きCSVで出力</td></tr>
            </table>
        </div>
        <div class="card" style="margin-top:0.5rem;">
            <div class="card-label">変換マップ（主要）</div>
            <table class="stat-table">
                <tr><th>元の値</th><th>変換後</th></tr>
                <tr><td>AI終話</td><td>NG</td></tr>
                <tr><td>応答なし</td><td>留守</td></tr>
                <tr><td>AI対応不可</td><td>留守電</td></tr>
                <tr><td>留守番電話</td><td>留守電</td></tr>
                <tr><td>転送　NG</td><td>NG</td></tr>
                <tr><td>転送　再コール</td><td>再コール・転送成功</td></tr>
                <tr><td>受電　電話APO</td><td>受電電話APO</td></tr>
                <tr><td>AIホットリード</td><td>再コール・転送成功</td></tr>
            </table>
        </div>
        """, unsafe_allow_html=True)


# ============================================================
# ページ: AI リスト成形
# ============================================================
def page_ai_list():
    st.markdown("""
    <div class="topbar-ai">
        <h2>AIテレアポ — リスト成形</h2>
        <p>FileMakerデータをAIテレアポ投入用CSVに変換（社名・電話番号・住所統合・IDを整形）</p>
    </div>
    """, unsafe_allow_html=True)

    col_main, col_side = st.columns([3, 1])

    with col_main:
        st.markdown('<div class="card"><div class="card-label"><span class="sn-ai">1</span>FileMakerデータ（Excel）をアップロード</div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("FileMakerから出力したExcelファイルを選択",
                                    type=['xlsx', 'xls'], key="ai_list_upload", label_visibility="collapsed")
        st.markdown('</div>', unsafe_allow_html=True)

        if uploaded:
            file_bytes = uploaded.read()
            st.markdown(f'<div class="alert-ok">読み込み完了 — {uploaded.name} ({len(file_bytes)//1024:,} KB)</div>', unsafe_allow_html=True)

            try:
                preview_df = pd.read_excel(BytesIO(file_bytes))
                st.markdown(f'<div class="alert-info">{len(preview_df):,} 件のデータを検出しました</div>', unsafe_allow_html=True)

                # ファイル名をコピーできるように表示
                st.markdown('<div class="card"><div class="card-label">読み込んだファイル名（クリックして全選択）</div>', unsafe_allow_html=True)
                st.code(uploaded.name, language=None)
                st.markdown('</div>', unsafe_allow_html=True)

                with st.expander("元データプレビュー（先頭5件）"):
                    st.dataframe(preview_df.head(5), use_container_width=True)

            except Exception as e:
                st.error(f"ファイル読み込みエラー: {e}")
                return

            st.markdown('<div class="card"><div class="card-label"><span class="sn-ai">2</span>成形を実行</div>', unsafe_allow_html=True)
            if st.button("成形を実行", type="primary", key="ai_list_run"):
                with st.spinner("処理中..."):
                    try:
                        csv_bytes, result_df, total_rows = process_ai_list(file_bytes)

                        st.markdown(f'<div class="alert-ok">成形完了 — {len(result_df):,} 件</div>', unsafe_allow_html=True)

                        job_id = add_job("AI リスト成形", uploaded.name, total_rows, len(result_df))
                        st.markdown(f'<div class="alert-info">ジョブID: <code>{job_id}</code></div>', unsafe_allow_html=True)

                        with st.expander("出力データプレビュー（先頭5件）"):
                            st.dataframe(result_df.head(5), use_container_width=True)

                        base = uploaded.name.rsplit('.', 1)[0]
                        date_str = datetime.datetime.now().strftime("%Y%m%d")
                        st.markdown("---")
                        st.download_button("CSVをダウンロード", data=csv_bytes,
                                           file_name=f"{base}_{date_str}_AIテレアポリスト.csv",
                                           mime="text/csv", key="ai_list_dl")
                    except Exception as e:
                        st.error(f"処理エラー: {e}")
            st.markdown('</div>', unsafe_allow_html=True)

    with col_side:
        st.markdown("""
        <div class="card">
            <div class="card-label">処理内容</div>
            <table class="stat-table">
                <tr><td>顧客名 → 社名 に列名変換</td></tr>
                <tr><td>社名を50文字でカット</td></tr>
                <tr><td>社名・電話番号・住所統合を抽出</td></tr>
                <tr><td>「IDの頭にID」→「ID」に変換して付与</td></tr>
                <tr><td>Shift-JIS / UTF-8 BOMで出力</td></tr>
            </table>
        </div>
        <div class="card" style="margin-top:0.5rem;">
            <div class="card-label">出力列</div>
            <table class="stat-table">
                <tr><th>#</th><th>列名</th></tr>
                <tr><td>1</td><td>社名（50文字）</td></tr>
                <tr><td>2</td><td>電話番号</td></tr>
                <tr><td>3</td><td>住所統合</td></tr>
                <tr><td>4</td><td>ID</td></tr>
            </table>
        </div>
        """, unsafe_allow_html=True)


# ============================================================
# ページ: AI FM返送
# ============================================================
def page_ai_fm():
    st.markdown("""
    <div class="topbar-ai">
        <h2>AIテレアポ — ファイルメーカー返送</h2>
        <p>AIテレアポの通話履歴CSVをアップロードして重複排除・列整形しExcel出力</p>
    </div>
    """, unsafe_allow_html=True)

    col_main, col_side = st.columns([3, 1])

    with col_main:
        st.markdown('<div class="card"><div class="card-label"><span class="sn-ai">1</span>AIテレアポ 通話履歴CSVをアップロード</div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("AIテレアポからエクスポートした通話履歴CSVを選択",
                                    type=['csv'], key="ai_fm_upload", label_visibility="collapsed")
        st.markdown('</div>', unsafe_allow_html=True)

        if uploaded:
            raw = uploaded.read()
            st.markdown(f'<div class="alert-ok">読み込み完了 — {uploaded.name} ({len(raw)//1024:,} KB)</div>', unsafe_allow_html=True)

            st.markdown('<div class="card"><div class="card-label"><span class="sn-ai">2</span>変換を実行</div>', unsafe_allow_html=True)
            if st.button("変換を実行", type="primary", key="ai_fm_run"):
                with st.spinner("処理中..."):
                    try:
                        excel_bytes, result_df, rows_before, rows_after, used_enc = process_ai_fm(raw, uploaded.name)

                        removed = rows_before - rows_after
                        st.markdown(
                            f'<div class="alert-ok">変換完了 — 入力: {rows_before:,} 件 / 出力: {rows_after:,} 件'
                            f' / 重複削除: {removed:,} 件 &nbsp;|&nbsp; 文字コード: {used_enc}</div>',
                            unsafe_allow_html=True
                        )

                        job_id = add_job("AI FM返送", uploaded.name, rows_before, rows_after)
                        st.markdown(f'<div class="alert-info">ジョブID: <code>{job_id}</code></div>', unsafe_allow_html=True)

                        if "ステータス" in result_df.columns:
                            with st.expander("ステータス分布"):
                                dist = result_df["ステータス"].value_counts().reset_index()
                                dist.columns = ["ステータス", "件数"]
                                st.dataframe(dist, use_container_width=True, hide_index=True)

                        with st.expander("出力データプレビュー（先頭5件）"):
                            st.dataframe(result_df.head(5), use_container_width=True)

                        base = uploaded.name.rsplit('.', 1)[0]
                        date_str = datetime.datetime.now().strftime("%Y%m%d")
                        st.markdown("---")
                        st.download_button(
                            "Excelをダウンロード", data=excel_bytes,
                            file_name=f"{base}_{date_str}_FM返送.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="ai_fm_dl"
                        )
                    except Exception as e:
                        st.error(f"処理エラー: {e}")
            st.markdown('</div>', unsafe_allow_html=True)

    with col_side:
        st.markdown("""
        <div class="card">
            <div class="card-label">処理内容</div>
            <table class="stat-table">
                <tr><td>IDの頭にID で重複排除（最新コールのみ残す）</td></tr>
                <tr><td>ソース列にアップロードファイル名を付与</td></tr>
                <tr><td>出力列を18列に固定</td></tr>
                <tr><td>1シートのExcel（.xlsx）で出力</td></tr>
            </table>
        </div>
        <div class="card" style="margin-top:0.5rem;">
            <div class="card-label">出力列（18列）</div>
            <table class="stat-table">
                <tr><th>#</th><th>列名</th></tr>
                <tr><td>1</td><td>日時</td></tr>
                <tr><td>2</td><td>種別</td></tr>
                <tr><td>3</td><td>会社名</td></tr>
                <tr><td>4</td><td>担当者名</td></tr>
                <tr><td>5</td><td>電話番号</td></tr>
                <tr><td>6</td><td>ステータス</td></tr>
                <tr><td>7</td><td>最終ステータス</td></tr>
                <tr><td>8</td><td>ラベル</td></tr>
                <tr><td>9</td><td>通話時間</td></tr>
                <tr><td>10</td><td>架電結果</td></tr>
                <tr><td>11</td><td>要約</td></tr>
                <tr><td>12</td><td>詳細リンク</td></tr>
                <tr><td>13</td><td>再電話回数</td></tr>
                <tr><td>14</td><td>メールアドレス</td></tr>
                <tr><td>15</td><td>最終試行</td></tr>
                <tr><td>16</td><td>総試行回数</td></tr>
                <tr><td>17</td><td>ソース</td></tr>
                <tr><td>18</td><td>IDの頭にID</td></tr>
            </table>
        </div>
        """, unsafe_allow_html=True)


# ============================================================
# ページ: ジョブ履歴
# ============================================================
def page_history():
    st.markdown("""
    <div class="topbar-hist">
        <h2>ジョブ履歴</h2>
        <p>処理済みジョブの一覧。定期に消さないと容量なくなる</p>
    </div>
    """, unsafe_allow_html=True)

    jobs = load_jobs()

    col_l, col_r = st.columns([4, 1])
    with col_l:
        st.markdown(f'<div class="alert-info">合計 {len(jobs):,} 件のジョブが記録されています</div>', unsafe_allow_html=True)
    with col_r:
        if st.button("履歴をクリア", key="clear_history"):
            save_jobs([])
            st.rerun()

    if not jobs:
        st.markdown("""
        <div style="text-align:center;padding:4rem 2rem;color:#94a3b8;">
            <p style="font-size:1rem;font-weight:600;">ジョブ履歴がありません</p>
            <p style="font-size:0.85rem;">各機能でファイルを処理するとここに記録されます</p>
        </div>
        """, unsafe_allow_html=True)
        return

    # フィルター
    filter_type = st.selectbox("種別フィルター", ["すべて", "DS リスト成形", "DS FM返送", "AI リスト成形", "AI FM返送"],
                                key="hist_filter", label_visibility="collapsed")

    filtered = jobs if filter_type == "すべて" else [j for j in jobs if j.get('type') == filter_type]

    for job in filtered:
        jtype = job.get('type', '')
        badge_class = "badge-ai" if jtype.startswith("AI") else "badge-ds"

        st.markdown(f"""
        <div class="job-card">
            <div class="job-card-header">
                <span class="job-id">{job.get('job_id', '-')}</span>
                <span class="{badge_class}">{jtype}</span>
            </div>
            <div class="job-meta">
                <div class="job-meta-item"><strong>処理日時</strong>　{job.get('created_at', '-')}</div>
                <div class="job-meta-item"><strong>入力</strong>　{job.get('rows_in', 0):,} 件</div>
                <div class="job-meta-item"><strong>出力</strong>　{job.get('rows_out', 0):,} 件</div>
            </div>
            <div class="job-fname">ファイル名: {job.get('filename', '-')}</div>
        </div>
        """, unsafe_allow_html=True)

    # CSVエクスポート
    st.markdown("---")
    hist_df = pd.DataFrame(filtered)
    hist_csv = hist_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
    st.download_button("履歴をCSVでダウンロード", data=hist_csv,
                       file_name=f"job_history_{datetime.datetime.now().strftime('%Y%m%d')}.csv",
                       mime="text/csv", key="hist_dl")


# ============================================================
# メイン
# ============================================================
def main():
    init_session()
    render_sidebar()

    page = st.session_state.get('page', 'home')
    if   page == 'home':    page_home()
    elif page == 'ds_list': page_ds_list()
    elif page == 'ds_fm':   page_ds_fm()
    elif page == 'ai_list': page_ai_list()
    elif page == 'ai_fm':   page_ai_fm()
    elif page == 'history': page_history()
    else:                   page_home()


if __name__ == "__main__":
    main()
